"""
CheckappExcel - Core comparator
Confronta due o più file Excel/CSV (anche con più fogli) usando il codice
prodotto come chiave primaria e genera un report Excel con colonne colorate.
"""

from __future__ import annotations

import os
from dataclasses import dataclass, field
from pathlib import Path
from typing import Dict, Iterable, List, Optional, Sequence, Tuple

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet


# ---------------------------------------------------------------------------
# Configurazione colonne
# ---------------------------------------------------------------------------

# Nomi "canonici" delle colonne che l'app cerca nei file di input.
# Per ciascun campo si elencano possibili varianti (case-insensitive, trimmed).
DEFAULT_COLUMN_ALIASES: Dict[str, List[str]] = {
    "codice": [
        "codice", "cod", "cod.", "cod prodotto", "codice prodotto",
        "codice articolo", "articolo", "sku", "id", "item code", "code",
    ],
    "descrizione": [
        "descrizione", "desc", "desc.", "descrizione prodotto",
        "denominazione", "nome", "description", "prodotto", "product",
        "descrizione tecnica", "descrizione commerciale", "descrizione breve",
        "nome breve", "nome commerciale", "nome prodotto", "nome articolo",
        "titolo", "item description", "product description",
    ],
    "prezzo": [
        "prezzo", "prezzo unitario", "prezzo netto", "prezzo listino",
        "importo", "costo", "price", "unit price",
        "prezzo lordo", "prezzolordo", "prezzo netto listino",
        "listino", "list price", "net price", "gross price",
    ],
    "trasporto": [
        "trasporto", "spese trasporto", "spedizione", "costo trasporto",
        "costotrasporto", "spese di trasporto", "shipping", "transport",
        "shipping cost", "trasporto e spedizione", "spese spedizione",
    ],
    "installazione": [
        "installazione", "montaggio", "installazione/montaggio",
        "costo installazione", "costoinstallazione", "costo montaggio",
        "spese installazione", "spese montaggio", "setup",
        "installation", "installation cost", "assembly",
    ],
}

# Campi comparati (oltre al codice che è la chiave).
COMPARISON_FIELDS: List[str] = ["descrizione", "prezzo", "trasporto", "installazione"]

# Palette di colori assegnati ciclicamente ai file di input (intestazioni).
FILE_HEADER_COLORS: List[str] = [
    "1F77B4",  # blu
    "2CA02C",  # verde
    "FF7F0E",  # arancio
    "9467BD",  # viola
    "17BECF",  # ciano
    "BCBD22",  # oliva
    "8C564B",  # marrone
    "E377C2",  # rosa
]

# Colori per le celle
COLOR_MISSING = "FFF2CC"        # giallo chiaro - prodotto assente in questo file
COLOR_PRESENT_ONLY = "C6EFCE"   # verde chiaro - presente solo qui
COLOR_PRICE_DIFF = "FCE4D6"     # arancio chiaro - prezzo diverso fra file
COLOR_HEADER_TEXT = "FFFFFF"    # testo bianco sulle intestazioni colorate
COLOR_KEY_FILL = "D9E1F2"       # azzurro tenue per colonna Codice
COLOR_STATUS_ALL = "C6EFCE"     # verde - in tutti i file
COLOR_STATUS_PARTIAL = "FFF2CC" # giallo - in alcuni
COLOR_STATUS_UNIQUE = "F8CBAD"  # arancio - in uno solo


# ---------------------------------------------------------------------------
# Modelli
# ---------------------------------------------------------------------------

@dataclass
class LoadedSource:
    """Rappresenta un file caricato e normalizzato."""
    label: str                          # etichetta mostrata nel report (es. "Fornitore_A")
    path: str                           # percorso file originale
    sheets: Dict[str, pd.DataFrame] = field(default_factory=dict)
    # chiave = nome foglio (o "" per CSV), valore = DataFrame normalizzato con
    # colonne canoniche ["codice","descrizione","prezzo","trasporto","installazione"]


@dataclass
class CompareOptions:
    """Opzioni configurabili per il confronto."""
    output_path: str = "confronto.xlsx"
    case_sensitive_codes: bool = False
    strip_codes: bool = True
    merge_sheets: bool = True           # se True, unisce tutti i fogli di un file in un'unica tabella
    column_aliases: Dict[str, List[str]] = field(default_factory=lambda: {
        k: list(v) for k, v in DEFAULT_COLUMN_ALIASES.items()
    })


# ---------------------------------------------------------------------------
# Utility
# ---------------------------------------------------------------------------

def _norm(s: str) -> str:
    """Normalizza un nome colonna.

    - lowercase, trim
    - underscore/trattini → spazio
    - spazi multipli → singolo
    - rimuove suffissi valuta (eur, euro, €, usd, $) dal finale
    - rimuove punteggiatura comune ai bordi
    """
    import re
    s = str(s).strip().lower()
    s = s.replace("_", " ").replace("-", " ").replace("/", " ")
    s = re.sub(r"\s+", " ", s).strip()
    # rimuovi suffissi valuta
    s = re.sub(r"\s*[\(\[]?\s*(eur|euro|€|usd|\$)\s*[\)\]]?\s*$", "", s).strip()
    # rimuovi punteggiatura estrema
    s = s.strip(".,:;")
    return s


def _build_alias_lookup(aliases: Dict[str, List[str]]) -> Dict[str, str]:
    """Mappa alias_normalizzato -> campo_canonico."""
    lookup: Dict[str, str] = {}
    for canonical, variants in aliases.items():
        for v in variants:
            lookup[_norm(v)] = canonical
        lookup[_norm(canonical)] = canonical
    return lookup


def _guess_columns(df: pd.DataFrame, alias_lookup: Dict[str, str]) -> Dict[str, str]:
    """
    Dato un DataFrame, restituisce un mapping
    {campo_canonico: nome_colonna_originale}.
    Se una colonna canonica non è presente, viene omessa.
    """
    result: Dict[str, str] = {}
    for col in df.columns:
        canon = alias_lookup.get(_norm(col))
        if canon and canon not in result:
            result[canon] = col
    return result


def _score_header_row(values: List, alias_lookup: Dict[str, str]) -> int:
    """Conta quante celle di una riga sembrano un header canonico."""
    canonical_seen = set()
    for v in values:
        if v is None:
            continue
        canon = alias_lookup.get(_norm(v))
        if canon:
            canonical_seen.add(canon)
    return len(canonical_seen)


def _detect_header_row(df: pd.DataFrame, alias_lookup: Dict[str, str],
                       max_scan: int = 10) -> Optional[int]:
    """Se l'header non è la riga 0, cerca quale delle prime max_scan
    righe sembra contenere i nomi colonna. Ritorna l'indice 0-based
    della riga, oppure None se non trovato."""
    # riga 0 (colonne correnti) come baseline
    best_row = -1  # -1 = colonne già OK
    best_score = _score_header_row(list(df.columns), alias_lookup)
    for i in range(min(max_scan, len(df))):
        score = _score_header_row(list(df.iloc[i].values), alias_lookup)
        if score > best_score:
            best_score = score
            best_row = i
    # serve almeno codice + 1 altro campo perché valga la pena promuoverla
    if best_row >= 0 and best_score >= 2:
        return best_row
    return None


def _normalize_code(value, *, case_sensitive: bool, strip: bool) -> Optional[str]:
    if value is None:
        return None
    if isinstance(value, float) and pd.isna(value):
        return None
    s = str(value)
    if strip:
        s = s.strip()
    if not s:
        return None
    # gestisce i codici numerici letti come float (es. 12345.0 -> "12345")
    try:
        f = float(s.replace(",", "."))
        if f.is_integer():
            s = str(int(f))
    except ValueError:
        pass
    if not case_sensitive:
        s = s.upper()
    return s


def _to_number(value) -> Optional[float]:
    """Prova a convertire in float (gestisce virgole e simboli €)."""
    if value is None:
        return None
    if isinstance(value, (int, float)):
        if pd.isna(value):
            return None
        return float(value)
    s = str(value).strip()
    if not s:
        return None
    s = s.replace("€", "").replace("$", "").replace(" ", "")
    s = s.replace(".", "") if s.count(",") == 1 and s.count(".") >= 1 else s
    s = s.replace(",", ".")
    try:
        return float(s)
    except ValueError:
        return None


# ---------------------------------------------------------------------------
# Caricamento file
# ---------------------------------------------------------------------------

def load_source(path: str, label: Optional[str] = None,
                options: Optional[CompareOptions] = None) -> LoadedSource:
    """Carica un file xlsx/xls/csv e lo normalizza."""
    options = options or CompareOptions()
    alias_lookup = _build_alias_lookup(options.column_aliases)

    p = Path(path)
    if not p.exists():
        raise FileNotFoundError(f"File non trovato: {path}")

    label = label or p.stem
    source = LoadedSource(label=label, path=str(p))

    ext = p.suffix.lower()
    if ext in (".csv", ".tsv", ".txt"):
        sep = "\t" if ext == ".tsv" else None  # None => pandas tenta di dedurre
        try:
            df = pd.read_csv(p, sep=sep, engine="python")
        except Exception:
            df = pd.read_csv(p, sep=";", engine="python")
        source.sheets[""] = _normalize_df(df, alias_lookup, options)
    elif ext in (".xlsx", ".xlsm", ".xls"):
        xl = pd.ExcelFile(p)
        for sheet_name in xl.sheet_names:
            df = xl.parse(sheet_name)
            if df.empty:
                continue
            norm = _normalize_df(df, alias_lookup, options)
            if norm is not None and not norm.empty:
                source.sheets[sheet_name] = norm
    else:
        raise ValueError(f"Formato non supportato: {ext}")

    if not source.sheets:
        raise ValueError(f"Nessun foglio utile trovato in {path} "
                         f"(verifica che ci sia almeno una colonna 'codice').")
    return source


def _normalize_df(df: pd.DataFrame, alias_lookup: Dict[str, str],
                  options: CompareOptions) -> Optional[pd.DataFrame]:
    """Trasforma un DataFrame nelle colonne canoniche."""
    # Auto-detect dell'header: se la riga 0 non ha "codice" ma qualche riga
    # successiva è chiaramente l'header, la promuoviamo.
    header_row = _detect_header_row(df, alias_lookup)
    if header_row is not None:
        new_columns = [str(v) if v is not None else "" for v in df.iloc[header_row].values]
        df = df.iloc[header_row + 1:].reset_index(drop=True).copy()
        df.columns = new_columns

    col_map = _guess_columns(df, alias_lookup)
    if "codice" not in col_map:
        return None  # foglio scartato: senza codice non possiamo confrontare

    out = pd.DataFrame()
    out["codice"] = df[col_map["codice"]].apply(
        lambda v: _normalize_code(
            v,
            case_sensitive=options.case_sensitive_codes,
            strip=options.strip_codes,
        )
    )

    for field_name in COMPARISON_FIELDS:
        if field_name in col_map:
            out[field_name] = df[col_map[field_name]]
        else:
            out[field_name] = pd.NA

    # Scarta righe senza codice
    out = out[out["codice"].notna() & (out["codice"].astype(str).str.len() > 0)]

    # In caso di duplicati nello stesso file teniamo la prima occorrenza
    out = out.drop_duplicates(subset=["codice"], keep="first").reset_index(drop=True)
    return out


# ---------------------------------------------------------------------------
# Confronto
# ---------------------------------------------------------------------------

def _flatten_source(source: LoadedSource, merge_sheets: bool
                    ) -> List[Tuple[str, pd.DataFrame]]:
    """Restituisce una lista di (label, df).

    Quando merge_sheets=True, i fogli dello stesso file vengono combinati in
    modo intelligente: per ogni codice, i valori vuoti in un foglio vengono
    riempiti con quelli trovati negli altri fogli. Questo evita di perdere
    prezzi/trasporto/installazione che sono distribuiti su fogli diversi.
    """
    if merge_sheets or len(source.sheets) == 1:
        frames = list(source.sheets.values())
        if len(frames) == 1:
            return [(source.label, frames[0].reset_index(drop=True))]

        # Combina i fogli riempiendo i vuoti
        combined: Dict[str, Dict] = {}  # codice -> record
        order: List[str] = []

        def _is_empty(v) -> bool:
            if v is None:
                return True
            if isinstance(v, float) and pd.isna(v):
                return True
            try:
                if pd.isna(v):
                    return True
            except (TypeError, ValueError):
                pass
            if isinstance(v, str) and v.strip() == "":
                return True
            return False

        for df in frames:
            for _, row in df.iterrows():
                code = row["codice"]
                if code not in combined:
                    combined[code] = {f: None for f in COMPARISON_FIELDS}
                    order.append(code)
                for f in COMPARISON_FIELDS:
                    if _is_empty(combined[code][f]):
                        v = row[f]
                        if _is_empty(v):
                            continue
                        combined[code][f] = v
        rows = []
        for code in order:
            rec = {"codice": code}
            rec.update(combined[code])
            rows.append(rec)
        merged = pd.DataFrame(rows)
        return [(source.label, merged.reset_index(drop=True))]
    return [(f"{source.label} / {sheet}", df)
            for sheet, df in source.sheets.items()]


def compare(sources: Sequence[LoadedSource],
            options: Optional[CompareOptions] = None) -> Dict:
    """
    Esegue il confronto fra sorgenti.
    Restituisce un dict con:
      - 'labels': lista delle etichette (colonne per ciascun file)
      - 'table': DataFrame affiancato con colonne per ogni file
      - 'stats': riepilogo conteggi
    """
    options = options or CompareOptions()
    if len(sources) < 2:
        raise ValueError("Servono almeno 2 file per fare un confronto.")

    # Espandi in liste piatte di (label, df)
    flat: List[Tuple[str, pd.DataFrame]] = []
    for s in sources:
        flat.extend(_flatten_source(s, options.merge_sheets))

    labels = [lbl for lbl, _ in flat]

    # Codici unione
    all_codes: List[str] = []
    seen = set()
    for _, df in flat:
        for c in df["codice"]:
            if c not in seen:
                seen.add(c)
                all_codes.append(c)

    # Costruisci tabella wide
    wide = pd.DataFrame({"codice": all_codes})
    presence_cols: List[str] = []
    for label, df in flat:
        sub = df.set_index("codice")
        for field_name in COMPARISON_FIELDS:
            col_name = f"{label} | {field_name}"
            wide[col_name] = wide["codice"].map(sub[field_name])
        presence_col = f"__present__{label}"
        wide[presence_col] = wide["codice"].isin(sub.index)
        presence_cols.append(presence_col)

    # Stato / status
    def _status(row) -> str:
        present = [row[c] for c in presence_cols]
        n_present = sum(present)
        if n_present == len(presence_cols):
            return "In tutti"
        if n_present == 1:
            idx = present.index(True)
            return f"Solo in {labels[idx]}"
        missing = [labels[i] for i, p in enumerate(present) if not p]
        return f"Manca in: {', '.join(missing)}"

    wide["Stato"] = wide.apply(_status, axis=1)
    # riordina: codice, Stato, poi blocchi per file
    ordered_cols = ["codice", "Stato"]
    for label in labels:
        for field_name in COMPARISON_FIELDS:
            ordered_cols.append(f"{label} | {field_name}")
    # Tieni anche le colonne di presenza per evidenziare in export
    final = wide[ordered_cols].copy()

    stats = {
        "totale_codici": len(final),
        "in_tutti": int((wide["Stato"] == "In tutti").sum()),
        "solo_in_uno": int(wide["Stato"].str.startswith("Solo in").sum()),
        "parziali": int(wide["Stato"].str.startswith("Manca in").sum()),
        "per_file": {
            label: int(wide[f"__present__{label}"].sum())
            for label in labels
        },
    }

    return {
        "labels": labels,
        "table": final,
        "presence": wide[["codice"] + presence_cols].rename(
            columns={f"__present__{l}": l for l in labels}
        ),
        "stats": stats,
        "options": options,
    }


# ---------------------------------------------------------------------------
# Export Excel con formattazione
# ---------------------------------------------------------------------------

def export_to_excel(result: Dict, output_path: Optional[str] = None) -> str:
    """Genera un file Excel formattato e ritorna il percorso prodotto."""
    options: CompareOptions = result["options"]
    output_path = output_path or options.output_path
    out = Path(output_path)
    out.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    _write_summary_sheet(wb.active, result)
    _write_compare_sheet(wb.create_sheet("Confronto"), result)
    _write_missing_sheet(wb.create_sheet("Mancanti"), result)

    wb.save(out)
    return str(out)


def _thin_border() -> Border:
    side = Side(style="thin", color="B7B7B7")
    return Border(left=side, right=side, top=side, bottom=side)


def _write_summary_sheet(ws: Worksheet, result: Dict) -> None:
    ws.title = "Riepilogo"
    stats = result["stats"]
    labels = result["labels"]

    ws["A1"] = "CheckappExcel - Riepilogo confronto"
    ws["A1"].font = Font(bold=True, size=14)

    rows: List[Tuple[str, object]] = [
        ("Totale codici distinti", stats["totale_codici"]),
        ("Presenti in tutti i file/fogli", stats["in_tutti"]),
        ("Presenti solo in uno", stats["solo_in_uno"]),
        ("Parzialmente presenti", stats["parziali"]),
    ]
    r = 3
    for k, v in rows:
        ws.cell(row=r, column=1, value=k).font = Font(bold=True)
        ws.cell(row=r, column=2, value=v)
        r += 1

    r += 1
    ws.cell(row=r, column=1, value="Conteggio per file/foglio").font = Font(bold=True)
    r += 1
    ws.cell(row=r, column=1, value="File / Foglio").font = Font(bold=True)
    ws.cell(row=r, column=2, value="N. codici").font = Font(bold=True)
    r += 1
    for i, label in enumerate(labels):
        color = FILE_HEADER_COLORS[i % len(FILE_HEADER_COLORS)]
        cell = ws.cell(row=r, column=1, value=label)
        cell.fill = PatternFill("solid", fgColor=color)
        cell.font = Font(color=COLOR_HEADER_TEXT, bold=True)
        ws.cell(row=r, column=2, value=stats["per_file"][label])
        r += 1

    # Legenda colori
    r += 2
    ws.cell(row=r, column=1, value="Legenda celle").font = Font(bold=True)
    r += 1
    legend = [
        ("Cella gialla = codice assente in quel file", COLOR_MISSING),
        ("Cella verde = codice presente solo in quel file", COLOR_PRESENT_ONLY),
        ("Cella arancio = prezzo diverso rispetto agli altri file", COLOR_PRICE_DIFF),
    ]
    for text, color in legend:
        c = ws.cell(row=r, column=1, value=text)
        c.fill = PatternFill("solid", fgColor=color)
        r += 1

    ws.column_dimensions["A"].width = 48
    ws.column_dimensions["B"].width = 20


def _compute_price_diff_flags(table: pd.DataFrame,
                              labels: List[str]) -> Dict[int, bool]:
    """Per ogni riga (indice 0-based), True se i prezzi numerici divergono."""
    flags: Dict[int, bool] = {}
    price_cols = [f"{l} | prezzo" for l in labels]
    for i, row in table.iterrows():
        nums = []
        for c in price_cols:
            n = _to_number(row.get(c))
            if n is not None:
                nums.append(round(n, 2))
        flags[i] = len(set(nums)) > 1
    return flags


def _write_compare_sheet(ws: Worksheet, result: Dict) -> None:
    table: pd.DataFrame = result["table"]
    labels: List[str] = result["labels"]
    presence: pd.DataFrame = result["presence"]  # colonna codice + booleani per label

    border = _thin_border()

    # --- Prima riga: titolo gruppo (nome file) con merge ---
    # Colonna 1: "Codice"
    # Colonna 2: "Stato"
    # Poi blocchi di 4 colonne (descrizione/prezzo/trasporto/installazione) per ogni label
    ws.cell(row=1, column=1, value="Codice")
    ws.cell(row=1, column=2, value="Stato")
    # Merge header per ogni file
    col = 3
    for i, label in enumerate(labels):
        color = FILE_HEADER_COLORS[i % len(FILE_HEADER_COLORS)]
        start_col = col
        end_col = col + len(COMPARISON_FIELDS) - 1
        ws.merge_cells(start_row=1, start_column=start_col,
                       end_row=1, end_column=end_col)
        cell = ws.cell(row=1, column=start_col, value=label)
        cell.font = Font(bold=True, color=COLOR_HEADER_TEXT, size=12)
        cell.fill = PatternFill("solid", fgColor=color)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        # Colora anche tutte le celle del merge (openpyxl lascia le altre senza fill)
        for c in range(start_col, end_col + 1):
            ws.cell(row=1, column=c).fill = PatternFill("solid", fgColor=color)
        col = end_col + 1

    # Intestazione riga 2 con nomi campi
    header2 = ["Codice", "Stato"]
    for _ in labels:
        header2.extend(["Descrizione", "Prezzo", "Trasporto", "Installazione"])
    for idx, name in enumerate(header2, start=1):
        c = ws.cell(row=2, column=idx, value=name)
        c.font = Font(bold=True)
        c.alignment = Alignment(horizontal="center", vertical="center",
                                wrap_text=True)
        c.border = border
        if idx <= 2:
            c.fill = PatternFill("solid", fgColor=COLOR_KEY_FILL)
        else:
            # riprendi il colore del gruppo ma più chiaro (usiamo il colore pieno
            # mantenendo coerenza visiva)
            block = (idx - 3) // len(COMPARISON_FIELDS)
            color = FILE_HEADER_COLORS[block % len(FILE_HEADER_COLORS)]
            c.fill = PatternFill("solid", fgColor=color)
            c.font = Font(bold=True, color=COLOR_HEADER_TEXT)

    # --- Dati ---
    price_diff_flags = _compute_price_diff_flags(table, labels)
    n_labels = len(labels)
    data_start_row = 3

    # Per velocità, pre-calcoliamo la presenza per riga/label
    presence_map: Dict[str, Dict[str, bool]] = {}
    for _, prow in presence.iterrows():
        presence_map[str(prow["codice"])] = {
            label: bool(prow[label]) for label in labels
        }

    for i, (_, row) in enumerate(table.iterrows()):
        r = data_start_row + i
        codice = row["codice"]
        ws.cell(row=r, column=1, value=codice).fill = PatternFill(
            "solid", fgColor=COLOR_KEY_FILL
        )
        stato = row["Stato"]
        status_cell = ws.cell(row=r, column=2, value=stato)
        if stato == "In tutti":
            status_cell.fill = PatternFill("solid", fgColor=COLOR_STATUS_ALL)
        elif stato.startswith("Solo in"):
            status_cell.fill = PatternFill("solid", fgColor=COLOR_STATUS_UNIQUE)
        else:
            status_cell.fill = PatternFill("solid", fgColor=COLOR_STATUS_PARTIAL)

        # Quanti file hanno il codice? (serve per il verde "presente solo qui")
        presents = presence_map.get(str(codice), {})
        n_present = sum(1 for v in presents.values() if v)

        col_cursor = 3
        for label in labels:
            is_present = presents.get(label, False)
            for field_name in COMPARISON_FIELDS:
                col_name = f"{label} | {field_name}"
                value = row.get(col_name)
                # NaN -> stringa vuota per pulizia
                if isinstance(value, float) and pd.isna(value):
                    value = None
                cell = ws.cell(row=r, column=col_cursor, value=value)
                cell.border = border
                cell.alignment = Alignment(vertical="center", wrap_text=True)

                if not is_present:
                    cell.fill = PatternFill("solid", fgColor=COLOR_MISSING)
                elif n_present == 1:
                    cell.fill = PatternFill("solid", fgColor=COLOR_PRESENT_ONLY)
                elif field_name == "prezzo" and price_diff_flags.get(i, False):
                    cell.fill = PatternFill("solid", fgColor=COLOR_PRICE_DIFF)

                # Format numerico per prezzo/trasporto/installazione
                if field_name in ("prezzo", "trasporto", "installazione"):
                    num = _to_number(value)
                    if num is not None:
                        cell.value = num
                        cell.number_format = '#,##0.00\\ "€"'
                col_cursor += 1

    # Larghezza colonne e freeze
    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 28
    for block_idx in range(n_labels):
        base = 3 + block_idx * len(COMPARISON_FIELDS)
        ws.column_dimensions[get_column_letter(base)].width = 40      # descrizione
        ws.column_dimensions[get_column_letter(base + 1)].width = 14  # prezzo
        ws.column_dimensions[get_column_letter(base + 2)].width = 14  # trasporto
        ws.column_dimensions[get_column_letter(base + 3)].width = 16  # installazione

    ws.freeze_panes = "C3"
    last_col = 2 + n_labels * len(COMPARISON_FIELDS)
    ws.auto_filter.ref = f"A2:{get_column_letter(last_col)}{ws.max_row}"


def _write_missing_sheet(ws: Worksheet, result: Dict) -> None:
    """Un foglio con solo i codici NON presenti in tutti i file."""
    table: pd.DataFrame = result["table"]
    missing = table[table["Stato"] != "In tutti"].copy()
    labels: List[str] = result["labels"]
    presence: pd.DataFrame = result["presence"]

    headers = ["Codice", "Stato"] + [f"In {l}?" for l in labels]
    for idx, h in enumerate(headers, start=1):
        c = ws.cell(row=1, column=idx, value=h)
        c.font = Font(bold=True, color=COLOR_HEADER_TEXT)
        c.fill = PatternFill("solid", fgColor="404040")
        c.alignment = Alignment(horizontal="center")

    pres_map = {str(row["codice"]): row for _, row in presence.iterrows()}

    for i, (_, row) in enumerate(missing.iterrows(), start=2):
        ws.cell(row=i, column=1, value=row["codice"]).fill = PatternFill(
            "solid", fgColor=COLOR_KEY_FILL
        )
        stato = row["Stato"]
        c = ws.cell(row=i, column=2, value=stato)
        if stato.startswith("Solo in"):
            c.fill = PatternFill("solid", fgColor=COLOR_STATUS_UNIQUE)
        else:
            c.fill = PatternFill("solid", fgColor=COLOR_STATUS_PARTIAL)
        pres_row = pres_map.get(str(row["codice"]))
        for j, label in enumerate(labels, start=3):
            has = bool(pres_row[label]) if pres_row is not None else False
            cc = ws.cell(row=i, column=j, value="Sì" if has else "No")
            cc.alignment = Alignment(horizontal="center")
            cc.fill = PatternFill(
                "solid", fgColor=COLOR_STATUS_ALL if has else COLOR_MISSING
            )

    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 32
    for j in range(3, 3 + len(labels)):
        ws.column_dimensions[get_column_letter(j)].width = 22
    ws.freeze_panes = "A2"
    if len(missing) > 0:
        ws.auto_filter.ref = (
            f"A1:{get_column_letter(2 + len(labels))}{ws.max_row}"
        )


# ---------------------------------------------------------------------------
# API di alto livello
# ---------------------------------------------------------------------------

def run_comparison(files: Iterable[str],
                   output_path: str = "confronto.xlsx",
                   labels: Optional[List[str]] = None,
                   options: Optional[CompareOptions] = None) -> Dict:
    """Carica i file, confronta ed esporta il risultato. Ritorna stats + path."""
    options = options or CompareOptions(output_path=output_path)
    options.output_path = output_path

    files = list(files)
    if labels is None:
        labels = [Path(f).stem for f in files]
    if len(labels) != len(files):
        raise ValueError("Il numero di etichette deve corrispondere al numero di file.")

    sources = [load_source(f, label=lbl, options=options)
               for f, lbl in zip(files, labels)]
    result = compare(sources, options=options)
    path = export_to_excel(result, output_path)
    return {"output": path, "stats": result["stats"], "labels": result["labels"]}
